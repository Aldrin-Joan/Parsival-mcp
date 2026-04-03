from __future__ import annotations
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from src.models.enums import FileFormat, ParseStatus
from src.models.metadata import DocumentMetadata


def _extract_shared_strings(workbook_zip: zipfile.ZipFile) -> list[str]:
    try:
        if "xl/sharedStrings.xml" not in workbook_zip.namelist():
            return []
        data = workbook_zip.read("xl/sharedStrings.xml")
        tree = ET.fromstring(data)
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        strings = []
        for si in tree.findall(".//s:si", ns):
            parts = [t.text or "" for t in si.findall(".//s:t", ns)]
            strings.append("".join(parts))
        return strings
    except Exception:
        return []


def _extract_sheets_from_workbook(workbook_zip: zipfile.ZipFile) -> list[tuple[str, str]]:
    sheets = []
    try:
        if "xl/workbook.xml" not in workbook_zip.namelist():
            return sheets

        workbook_xml = workbook_zip.read("xl/workbook.xml")
        tree = ET.fromstring(workbook_xml)
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        sheet_names = []
        for sheet in tree.findall(".//s:sheet", ns):
            name = sheet.attrib.get("name")
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if name and rel_id:
                sheet_names.append((name, rel_id))

        rels = {}
        if "xl/_rels/workbook.xml.rels" in workbook_zip.namelist():
            rel_tree = ET.fromstring(workbook_zip.read("xl/_rels/workbook.xml.rels"))
            for rel in rel_tree.findall(".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rid and target:
                    rels[rid] = target

        for name, rel_id in sheet_names:
            target = rels.get(rel_id)
            if target:
                sheets.append((name, target))

    except Exception:
        return []

    return sheets
from src.models.parse_result import ParseResult, ParseError, Section
from src.models.table import TableResult, TableCell
from src.parsers.base import BaseParser
from src.parsers.registry import register
from src.core.logging import get_logger

logger = get_logger(__name__)


@register(FileFormat.XLSX)
class XlsxParser(BaseParser):
    async def parse(self, path: Path, options: dict | None = None) -> ParseResult:
        src = Path(path)
        start = __import__("time").time()

        try:
            size_bytes = src.stat().st_size
            if size_bytes > (__import__("sys").maxsize):
                raise OverflowError("file too large")

            max_size = (options or {}).get("max_size_mb", None)
            if max_size is not None and size_bytes > max_size * 1024 * 1024:
                metadata = DocumentMetadata(
                    source_path=str(src),
                    file_format=FileFormat.XLSX,
                    file_size_bytes=size_bytes,
                    section_count=0,
                    table_count=0,
                    image_count=0,
                    has_toc=False,
                )
                return ParseResult(
                    status=ParseStatus.OVERSIZE,
                    metadata=metadata,
                    sections=[],
                    images=[],
                    tables=[],
                    errors=[ParseError(code="oversize", message="File exceeds configured max size", recoverable=False)],
                    raw_text=None,
                    cache_hit=False,
                    request_id="",
                )

            if size_bytes > 10 * 1024 * 1024:
                wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
                read_only_mode = True
            else:
                wb = openpyxl.load_workbook(str(src), read_only=False, data_only=True)
                read_only_mode = False
        except (InvalidFileException, Exception) as exc:
            # Attempt a fallback from raw ZIP contents for damaged or cropped XLSX
            try:
                with zipfile.ZipFile(src, "r") as z:
                    shared = _extract_shared_strings(z)
                    sheets = _extract_sheets_from_workbook(z)
                    tables = []
                    table_index = 0
                    for sheet_idx, (sheet_name, target) in enumerate(sheets):
                        sheet_path = "xl/" + target.lstrip("/")
                        if sheet_path not in z.namelist():
                            continue
                        sheet_xml = ET.fromstring(z.read(sheet_path))
                        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
                        rows_data = []
                        for row in sheet_xml.findall(".//s:row", ns):
                            row_values = []
                            for c in row.findall(".//s:c", ns):
                                value = ""
                                v = c.find("s:v", ns)
                                if v is not None and v.text is not None:
                                    if c.attrib.get("t") == "s":
                                        idx = int(v.text) if v.text.isdigit() else None
                                        value = shared[idx] if idx is not None and idx < len(shared) else v.text
                                    else:
                                        value = v.text
                                row_values.append(value)
                            rows_data.append(row_values)

                        if not rows_data:
                            continue

                        headers = rows_data[0]
                        body_rows = rows_data[1:]
                        cells = []
                        for ri, row in enumerate(rows_data):
                            for ci, val in enumerate(row):
                                cells.append(
                                    TableCell(
                                        row=ri,
                                        col=ci,
                                        value=str(val),
                                        raw_value=val,
                                        colspan=1,
                                        rowspan=1,
                                        is_header=(ri == 0),
                                    )
                                )

                        tables.append(
                            TableResult(
                                index=table_index,
                                page=sheet_idx + 1,
                                caption=sheet_name,
                                headers=headers,
                                rows=body_rows,
                                cells=cells,
                                row_count=len(body_rows),
                                col_count=len(headers),
                                has_merged_cells=False,
                                confidence=0.7,
                                confidence_reason="xlsx fallback parser",
                                markdown="",
                                errors=["fallback_recover"],
                            )
                        )
                        table_index += 1

                    total_text = []
                    for table in tables:
                        total_text.extend([" ".join(r) for r in table.rows])

                    metadata = DocumentMetadata(
                        source_path=str(src),
                        file_format=FileFormat.XLSX,
                        file_size_bytes=src.stat().st_size if src.exists() else 0,
                        page_count=len(sheets),
                        word_count=len(" ".join(total_text).split()),
                        char_count=len(" ".join(total_text)),
                        reading_time_minutes=None,
                        section_count=0,
                        table_count=len(tables),
                        image_count=0,
                        has_toc=False,
                        toc=[],
                        parse_duration_ms=(__import__("time").time() - start) * 1000,
                        parser_version="xlsx_fallback",
                    )

                    return ParseResult(
                        status=ParseStatus.PARTIAL if tables else ParseStatus.FAILED,
                        metadata=metadata,
                        sections=[],
                        images=[],
                        tables=tables,
                        errors=[ParseError(code="corrupt_xlsx", message=str(exc), recoverable=True)],
                        raw_text="\n".join([" ".join(r) for t in tables for r in t.rows]) if tables else "",
                        cache_hit=False,
                        request_id="",
                    )
            except Exception:
                pass

            metadata = DocumentMetadata(
                source_path=str(src),
                file_format=FileFormat.XLSX,
                file_size_bytes=src.stat().st_size if src.exists() else 0,
                section_count=0,
                table_count=0,
                image_count=0,
                has_toc=False,
            )
            return ParseResult(
                status=ParseStatus.FAILED,
                metadata=metadata,
                sections=[],
                images=[],
                tables=[],
                errors=[ParseError(code="corrupt_xlsx", message=str(exc), recoverable=False)],
                raw_text="",
                cache_hit=False,
                request_id="",
            )

        sections: list[Section] = []
        tables: list[TableResult] = []
        errors: list[ParseError] = []

        table_index = 0
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            try:
                ws = wb[sheet_name]
                rows = []
                merged_map = {}
                for merged in ws.merged_cells.ranges:
                    merged_cells = ws[merged.coord]
                    for row_cells in merged_cells:
                        if isinstance(row_cells, tuple):
                            iter_cells = row_cells
                        else:
                            iter_cells = (row_cells,)
                        for cell in iter_cells:
                            if hasattr(cell, "coordinate"):
                                merged_map[cell.coordinate] = cell

                for row in ws.iter_rows(values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])

                if not rows:
                    continue

                headers = rows[0]
                body_rows = rows[1:]
                cells = []
                for ri, row in enumerate(rows):
                    for ci, val in enumerate(row):
                        cells.append(
                            TableCell(
                                row=ri,
                                col=ci,
                                value=str(val),
                                raw_value=val,
                                colspan=1,
                                rowspan=1,
                                is_header=(ri == 0),
                            )
                        )

                table_obj = TableResult(
                    index=table_index,
                    page=sheet_idx + 1,
                    caption=sheet_name,
                    headers=headers,
                    rows=body_rows,
                    cells=cells,
                    row_count=len(body_rows),
                    col_count=len(headers),
                    has_merged_cells=bool(ws.merged_cells),
                    confidence=0.9,
                    confidence_reason="openpyxl xlsx",
                    markdown="",
                    errors=[],
                )
                tables.append(table_obj)
                table_index += 1
            except Exception as exc:
                logger.warning("xlsx_sheet_parse_skipped", sheet=sheet_name, error=str(exc))
                errors.append(
                    ParseError(
                        code="xlsx_sheet_parse_error",
                        message=f"Sheet '{sheet_name}' skipped: {exc}",
                        recoverable=True,
                    )
                )
                continue

        total_text = []
        for table in tables:
            total_text.extend([" ".join(r) for r in table.rows])

        metadata = DocumentMetadata(
            source_path=str(src),
            file_format=FileFormat.XLSX,
            file_size_bytes=src.stat().st_size,
            page_count=len(wb.sheetnames),
            word_count=len(" ".join(total_text).split()),
            char_count=len(" ".join(total_text)),
            reading_time_minutes=None,
            section_count=len(sections),
            table_count=len(tables),
            image_count=0,
            has_toc=False,
            toc=[],
            parse_duration_ms=(__import__("time").time() - start) * 1000,
            parser_version=openpyxl.__version__,
        )

        wb.close()
        return ParseResult(
            status=ParseStatus.OK if tables and not errors else (ParseStatus.PARTIAL if tables else ParseStatus.FAILED),
            metadata=metadata,
            sections=sections,
            images=[],
            tables=tables,
            errors=errors,
            raw_text="\n".join([" ".join(r) for table in tables for r in table.rows]) if tables else "",
            cache_hit=False,
            request_id="",
        )

    async def parse_metadata(self, path: Path) -> DocumentMetadata:
        src = Path(path)
        try:
            wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)
            props = wb.properties

            metadata = DocumentMetadata(
                source_path=str(src),
                file_format=FileFormat.XLSX,
                file_size_bytes=src.stat().st_size,
                title=props.title,
                author=props.creator,
                subject=props.subject,
                keywords=props.keywords.split(",") if props.keywords else [],
                created_at=props.created.isoformat() if props.created else None,
                modified_at=props.modified.isoformat() if props.modified else None,
                producer=None,
                page_count=len(wb.sheetnames),
                section_count=0,
                table_count=0,
                image_count=0,
                has_toc=False,
                toc=[],
                parse_duration_ms=0.0,
                parser_version=openpyxl.__version__,
            )
            wb.close()
            return metadata
        except Exception:
            return DocumentMetadata(
                source_path=str(src),
                file_format=FileFormat.XLSX,
                file_size_bytes=src.stat().st_size if src.exists() else 0,
                title=None,
                author=None,
                subject=None,
                keywords=[],
                created_at=None,
                modified_at=None,
                producer=None,
                page_count=None,
                section_count=0,
                table_count=0,
                image_count=0,
                has_toc=False,
                toc=[],
                parse_duration_ms=0.0,
                parser_version="xlsx_parser_corrupt",
            )
