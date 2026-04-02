import os
import tempfile
from pathlib import Path
import openpyxl
from src.parsers.xlsx_parser import XlsxParser
from src.parsers.csv_parser import CsvParser
from src.models.enums import ParseStatus


def make_xlsx_file():
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "A"
    ws["B1"] = "B"
    ws["A2"] = "1"
    ws["B2"] = "2"
    wb.save(tmp.name)
    return tmp.name


def test_xlsx_parse_and_metadata():
    path = make_xlsx_file()
    try:
        parser = XlsxParser()
        result = __import__("asyncio").run(parser.parse(Path(path)))
        assert result.status == ParseStatus.OK
        assert result.metadata.file_format == "xlsx"
        assert result.metadata.table_count == 1
        assert len(result.tables) == 1

        md = __import__("asyncio").run(parser.parse_metadata(Path(path)))
        assert md.file_format == "xlsx"
        assert md.page_count == 1
    finally:
        os.unlink(path)


def make_csv_file():
    tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
    tmp.write(b"A,B\n1,2\n")
    tmp.close()
    return tmp.name


def test_csv_parse_and_metadata():
    path = make_csv_file()
    try:
        parser = CsvParser()
        result = __import__("asyncio").run(parser.parse(Path(path)))
        assert result.status == ParseStatus.OK
        assert result.metadata.file_format == "csv"
        assert result.metadata.table_count == 1
        assert len(result.tables) == 1

        md = __import__("asyncio").run(parser.parse_metadata(Path(path)))
        assert md.file_format == "csv"
        assert md.table_count == 1
    finally:
        os.unlink(path)
